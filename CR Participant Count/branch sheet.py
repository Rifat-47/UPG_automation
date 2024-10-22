"""update data on excel sheet"""

import sys
from openpyxl import load_workbook

# Your data (list of dictionaries)
branches_id = [
"4f719106-0a56-4ea0-8c4b-c69a94521e5e",
"5fd239e9-4006-4d66-a014-d8bac244f781",
"5373f8d1-23cf-4f5a-9e69-4ddcef4439a2",
"00423b68-c1d2-4c13-847e-374c1dbc6885",
"75611b6b-6a57-447e-8dcd-054bc14a1549",
"b5fec398-82da-457c-8ad6-b41446710178",
"be5fe4c6-a58c-4562-aa40-ba29d0aae7de",
"cf1550f1-6c09-471b-957f-3273e9e08da1",
"95badf44-4627-4d56-924b-0e291ab4515d",
"bbcefc16-d05f-4810-ab59-30c0fccdbaa1",
"0de10162-a694-428e-840f-3bdd3489b7ad",
"439beef0-4688-4fd6-addb-3e7fd4b4c6fa",
"3d074214-0621-4c89-a089-0df1c983b737",
"0323a12e-a12e-49b8-91a4-bce0ea7b7fc0",
]

abc = [
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":371,
      "group_visit_april_to_june":357,
      "home_visit_jan_to_march":363,
      "home_visit_april_to_june":374,
      "gvHv_jan_to_march":361,
      "gvHv_april_to_june":357,
      "group_visit_jan_to_march_by_question":229,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":173,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":8,
      "home_visit_april_to_june_answer_by_question":19
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":194,
      "group_visit_april_to_june":271,
      "home_visit_jan_to_march":218,
      "home_visit_april_to_june":275,
      "gvHv_jan_to_march":181,
      "gvHv_april_to_june":271,
      "group_visit_jan_to_march_by_question":138,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":119,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":2,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":303,
      "group_visit_april_to_june":304,
      "home_visit_jan_to_march":302,
      "home_visit_april_to_june":299,
      "gvHv_jan_to_march":301,
      "gvHv_april_to_june":298,
      "group_visit_jan_to_march_by_question":238,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":208,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":2,
      "home_visit_april_to_june_answer_by_question":3
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":303,
      "group_visit_april_to_june":303,
      "home_visit_jan_to_march":303,
      "home_visit_april_to_june":303,
      "gvHv_jan_to_march":303,
      "gvHv_april_to_june":303,
      "group_visit_jan_to_march_by_question":286,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":231,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":0,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":330,
      "group_visit_april_to_june":330,
      "home_visit_jan_to_march":330,
      "home_visit_april_to_june":330,
      "gvHv_jan_to_march":330,
      "gvHv_april_to_june":330,
      "group_visit_jan_to_march_by_question":305,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":228,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":22,
      "home_visit_april_to_june_answer_by_question":1
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":305,
      "group_visit_april_to_june":305,
      "home_visit_jan_to_march":305,
      "home_visit_april_to_june":305,
      "gvHv_jan_to_march":305,
      "gvHv_april_to_june":305,
      "group_visit_jan_to_march_by_question":249,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":174,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":0,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":306,
      "group_visit_april_to_june":310,
      "home_visit_jan_to_march":308,
      "home_visit_april_to_june":310,
      "gvHv_jan_to_march":304,
      "gvHv_april_to_june":310,
      "group_visit_jan_to_march_by_question":292,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":204,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":3,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":240,
      "group_visit_april_to_june":158,
      "home_visit_jan_to_march":321,
      "home_visit_april_to_june":314,
      "gvHv_jan_to_march":240,
      "gvHv_april_to_june":158,
      "group_visit_jan_to_march_by_question":111,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":174,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":0,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":328,
      "group_visit_april_to_june":305,
      "home_visit_jan_to_march":326,
      "home_visit_april_to_june":327,
      "gvHv_jan_to_march":326,
      "gvHv_april_to_june":304,
      "group_visit_jan_to_march_by_question":278,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":199,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":0,
      "home_visit_april_to_june_answer_by_question":4
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":312,
      "group_visit_april_to_june":312,
      "home_visit_jan_to_march":312,
      "home_visit_april_to_june":312,
      "gvHv_jan_to_march":312,
      "gvHv_april_to_june":312,
      "group_visit_jan_to_march_by_question":295,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":229,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":0,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":311,
      "group_visit_april_to_june":311,
      "home_visit_jan_to_march":312,
      "home_visit_april_to_june":312,
      "gvHv_jan_to_march":311,
      "gvHv_april_to_june":311,
      "group_visit_jan_to_march_by_question":240,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":235,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":1,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":322,
      "group_visit_april_to_june":321,
      "home_visit_jan_to_march":322,
      "home_visit_april_to_june":322,
      "gvHv_jan_to_march":322,
      "gvHv_april_to_june":321,
      "group_visit_jan_to_march_by_question":296,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":238,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":19,
      "home_visit_april_to_june_answer_by_question":1
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":326,
      "group_visit_april_to_june":324,
      "home_visit_jan_to_march":327,
      "home_visit_april_to_june":327,
      "gvHv_jan_to_march":326,
      "gvHv_april_to_june":324,
      "group_visit_jan_to_march_by_question":263,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":261,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":30,
      "home_visit_april_to_june_answer_by_question":0
   }
},
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":324,
      "group_visit_april_to_june":305,
      "home_visit_jan_to_march":328,
      "home_visit_april_to_june":345,
      "gvHv_jan_to_march":308,
      "gvHv_april_to_june":305,
      "group_visit_jan_to_march_by_question":297,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":230,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":2,
      "home_visit_april_to_june_answer_by_question":4
   }
},
]

if len(abc) != len(branches_id):
    sys.exit(10)

# Update each entry in abc to include branch_id at the start of the result dictionary
for entry, branch_id in zip(abc, branches_id):
    # Create a new result dictionary with branch_id as the first key
    new_result = {'branch_id': branch_id}
    new_result.update(entry["result"])

    # Update the entry with the new result dictionary
    entry["result"] = new_result

# Extract headers dynamically from the keys of the 'result' dictionary of the first entry
headers = ['', *abc[0]["result"].keys()]

# Specify the Excel file path
file_path = "UPG Rural 2023.xlsx"

# Load the workbook
workbook = load_workbook(file_path)
sheet = workbook.active

# Check if the first row is empty; if so, add the dynamic header
if sheet.max_row == 1 and not any(sheet["A1"].value for cell in sheet[1]):
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

# Append each entry in the list to the Excel sheet
for entry in abc:
    # Extract the data for each entry
    data = list(entry["result"].values())
    # Insert a blank value at the beginning of the data (for the first column)
    data.insert(0, None)
    # Find the next available row
    next_row = sheet.max_row + 1
    # Append the data to the next available row
    for col_num, value in enumerate(data, 1):
        sheet.cell(row=next_row, column=col_num, value=value)

# Save the workbook (this will save changes even if the file is open)
workbook.save(file_path)
