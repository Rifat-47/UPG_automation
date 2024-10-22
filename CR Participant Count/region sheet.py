"""update data on excel sheet"""

import sys
from openpyxl import load_workbook

# Your data (list of dictionaries)
regions_id = [
"33adb7e7-7f65-4487-99cc-40f2db4b66d2",
]

abc = [
{
   "status":"ok",
   "result":{
      "group_visit_jan_to_march":1777,
      "group_visit_april_to_june":1771,
      "home_visit_jan_to_march":1717,
      "home_visit_april_to_june":1858,
      "gvHv_jan_to_march":1668,
      "gvHv_april_to_june":1765,
      "group_visit_jan_to_march_by_question":1549,
      "group_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_by_question":1011,
      "home_visit_april_to_june_by_question":0,
      "home_visit_jan_to_march_answer_by_question":6,
      "home_visit_april_to_june_answer_by_question":34
   }
},

]

if len(abc) != len(regions_id):
    sys.exit(10)

# Update each entry in abc to include branch_id at the start of the result dictionary
for entry, branch_id in zip(abc, regions_id):
    # Create a new result dictionary with branch_id as the first key
    new_result = {'branch_id': branch_id}
    new_result.update(entry["result"])

    # Update the entry with the new result dictionary
    entry["result"] = new_result

# Extract headers dynamically from the keys of the 'result' dictionary of the first entry
headers = ['', *abc[0]["result"].keys()]

# Specify the Excel file path
file_path = "UPG Rural 2023.xlsx"

# Load the workbook
workbook = load_workbook(file_path)
sheet = workbook.active

# Check if the first row is empty; if so, add the dynamic header
if sheet.max_row == 1 and not any(sheet["A1"].value for cell in sheet[1]):
    for col_num, header in enumerate(headers, 1):
        sheet.cell(row=1, column=col_num, value=header)

# Append each entry in the list to the Excel sheet
for entry in abc:
    # Extract the data for each entry
    data = list(entry["result"].values())
    # Insert a blank value at the beginning of the data (for the first column)
    data.insert(0, None)
    # Find the next available row
    next_row = sheet.max_row + 1
    # Append the data to the next available row
    for col_num, value in enumerate(data, 1):
        sheet.cell(row=next_row, column=col_num, value=value)

# Save the workbook (this will save changes even if the file is open)
workbook.save(file_path)
