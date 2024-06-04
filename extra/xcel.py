import openpyxl
import os

def get_excel_files(folder_path):
    # List all files in the specified folder and filter for .xlsx files
    return [os.path.join(folder_path, f) for f in os.listdir(folder_path) if f.endswith('.xlsx')]

def merge_sheets(file_list, sheet_name):
    merged_data = []
    first_file = True

    for file in file_list:
        wb = openpyxl.load_workbook(file, data_only=True)
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            # Iterate over rows, skip the first row
            for row in sheet.iter_rows(values_only=True):
                if first_file:
                    # Keep the first row from the first file (column names)
                    merged_data.append(row)
                    first_file = False
                else:
                    # Skip the first row from subsequent files
                    merged_data.append(row)
            first_file = False

    # Remove the first row after merging
    return merged_data[1:]

def write_to_new_workbook(merged_data_dict, output_file):
    new_wb = openpyxl.Workbook()
    for sheet_name, data in merged_data_dict.items():
        new_ws = new_wb.create_sheet(title=sheet_name)
        for row in data:
            new_ws.append(row)
    if 'Sheet' in new_wb.sheetnames:
        del new_wb['Sheet']
    new_wb.save(output_file)

def get_sheet_names(file):
    # Load the workbook and return the sheet names
    wb = openpyxl.load_workbook(file, data_only=True)
    return wb.sheetnames

def main():
    # Specify the folder containing the Excel files
    folder_path = 'E:/BARIND_2024_BOTTOM_THREE/BARIND_2024_BOTTOM_THREE'
    file_list = get_excel_files(folder_path)
    # List of sheet names to merge
    sheet_names = get_sheet_names(file_list[0])

    merged_data_dict = {}

    for sheet_name in sheet_names:
        merged_data_dict[sheet_name] = merge_sheets(file_list, sheet_name)

    # Output file name
    output_file = 'BARIND_2024_BOTTOM_THREE.xlsx'
    write_to_new_workbook(merged_data_dict, output_file)

if __name__ == "__main__":
    main()
