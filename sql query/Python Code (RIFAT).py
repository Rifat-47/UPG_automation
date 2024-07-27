"""Read Carefully"""
"""it wil take 2-5 minutes to execute the code, sit tight, have some patience!"""
"""this code can browse & upload any xl sheet. In the code, first 5 columns are used for 'Variable'.
5th column is used for 'Time' & 6th column is used for 'Values'."""
"""after running the code, a new sheet named "Output" will be created in your excel file, where you can see all result 
for each variable. good luck!!!"""

import tkinter as tk
from tkinter import filedialog
import pandas as pd
import openpyxl

# Create a Tkinter window to select the file
root = tk.Tk()
root.withdraw()

file_path = filedialog.askopenfilename()

variableArray = []
finalResult = []

try:
    # Load the Excel sheet into a pandas dataframe
    df = pd.read_excel(file_path)
    # Check if the dataframe is empty
    if df.empty:
        print("The Excel sheet is empty. Please give a excel sheet with data")
        del df
    else:
        # Iterate over the rows of the dataframe using a for loop
        for row in range(len(df)):
            # Making columns for variables
            columns = df.columns[:5].tolist()
            # Iterate over the columns on each row of the dataframe using a for loop
            for column in columns:
                if not df.iloc[row][column] in variableArray:
                    if pd.isnull(df.loc[row, column]) or len(str(df.loc[row, column]).strip()) == 0:
                        continue
                    else:
                        variableArray.append(df.iloc[row][column])
                        dictionary = {'variable': df.iloc[row][column], 'length': 1, 'maxTime': df.iloc[row][df.columns[5]],
                            'minTime': df.iloc[row][df.columns[5]], 'totalValue': df.iloc[row][df.columns[6]],
                            'average': df.iloc[row][df.columns[6]]}
                        finalResult.append(dictionary)
                else:
                    for x in finalResult:
                        if x['variable'] == df.iloc[row][column]:
                            x['length'] += 1
                            # checking if current time is greater  than current maximum time
                            if df.iloc[row][df.columns[5]] > x['maxTime']:
                                x['maxTime'] = df.iloc[row][df.columns[5]]
                            # checking if current time is smaller than current minimum time
                            if df.iloc[row][df.columns[5]] < x['minTime']:
                                x['minTime'] = df.iloc[row][df.columns[5]]
                            x['totalValue'] += df.iloc[row][df.columns[6]]
                            x['average'] = x['totalValue']/x['length']
                        else:
                            continue
        del df

        # load the existing Excel file using openpyxl
        workbook = openpyxl.load_workbook(file_path)

        # create the sheet
        new_sheet = workbook.create_sheet("Output")

        # column names for output data
        result_columns = ["Variable", "Max Time", "Min Time", "Average"]
        for index, value in enumerate(result_columns):
            new_sheet.cell(row=1, column=index+1).value = value
        # new_sheet.cell(row=1, column=1).value = "Variable"
        # new_sheet.cell(row=1, column=2).value = "Max Time"
        # new_sheet.cell(row=1, column=3).value = "Min Time"
        # new_sheet.cell(row=1, column=4).value = "Average"

        # updating the output in the new sheet
        index = 0
        for row in range(2, len(finalResult) + 2):
            new_sheet.cell(row=row, column=1).value = finalResult[index].get('variable')
            new_sheet.cell(row=row, column=2).value = finalResult[index].get('maxTime')
            new_sheet.cell(row=row, column=3).value = finalResult[index].get('minTime')
            new_sheet.cell(row=row, column=4).value = finalResult[index].get('average')
            index += 1

        # save the workbook
        workbook.save(file_path)

        # printing output in the terminal
        for data in finalResult:
            print({'variable': data['variable'], 'maxTime': data['maxTime'], 'minTime': data['minTime'],
                'averageValue': data['average']})
                
except Exception as e:
    print("An error occured: ", e)