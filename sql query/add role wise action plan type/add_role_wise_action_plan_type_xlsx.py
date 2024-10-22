import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('../Book1.xlsx')

# Get all sheet names
sheet_names = workbook.sheetnames

all_data = []
query_list = []

for sheet_name in sheet_names:
    # Select the sheet by name
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column):
        row_data = [cell.value for cell in row]
        all_data.append(row_data)
    break

print(all_data)

for data in all_data:
    query_list.append(
        f'INSERT INTO role_wise_action_plan_type ('
        f'"cohort_id","role_id","action_plan_type_id","action_plan_type_name","created_at","created_by","is_active",'
        f'"role_name","updated_at","updated_by") VALUES ('
        f"{data[0]}, "
        f"{data[1]}, {data[2]}, '{data[3]}', toTimestamp(now()), '{data[5]}', {data[6]}, '{data[7]}', null, null);"
    )

for query in query_list:
    print(query)

print(len(query_list))