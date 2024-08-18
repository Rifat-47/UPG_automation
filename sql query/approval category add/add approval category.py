# INSERT INTO approval_category
# ("cohort_id","category_key","id","attribute","category_key_notification","category_name","created_at",
# "created_by","is_active","is_adhoc_approval","updated_at","updated_by") VALUES
# ('df90c5c9-7fb3-4633-99db-259b2ce82990','APPROVAL_SECOND_ROUND_SUPPORT',150d3037bceb-4c38-8a22-58191d485c66,
# NULL,'second_round_support_request.created','Second Round Asset Support','2022-11-10 07:27:33.000+0000',
# '47348703-ee43-4012-a3a1-1a5ffdef7c4e','true','false','2022-11-10 07:27:33.000+0000',NULL);

"""for setting approval hierarchy, enter parent data in excel sheet and run this script, then add
approval hierarchy configuration"""

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('../Book1.xlsx')

# Get all sheet names
sheet_names = workbook.sheetnames

all_data = []
query_list = []

for sheet_name in sheet_names:
    # Select the sheet by name
    sheet = workbook[sheet_name]

    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column):
        row_data = [cell.value for cell in row]
        all_data.append(row_data)
    break

print(all_data)

for data in all_data:
    query_list.append(
        f'INSERT INTO approval_category ('
        f'"cohort_id","category_key","id","attribute","category_key_notification","category_name","created_at",'
        f'"created_by","is_active","is_adhoc_approval","updated_at","updated_by") VALUES ('
        f"'6a1325d4-07cc-4985-b7cc-c945c4f536fe', "
        f"'{data[1]}', uuid(), null, '{data[4]}', '{data[5]}', '{data[6]}', '{data[7]}', true, true, '{data[10]}', null);"
    )

for query in query_list:
    print(query)

print(len(query_list))

