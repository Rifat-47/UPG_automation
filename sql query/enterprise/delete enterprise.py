import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('../Book1.xlsx')

# Get all sheet names
sheet_names = workbook.sheetnames

all_data = []
query_list = []

for sheet_name in sheet_names:
    # Select the sheet by name
    sheet = workbook[sheet_name]
    # print(sheet)
    # print(sheet.min_row, sheet.max_row)
    # print(sheet.min_column, sheet.max_column)

    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column):
        # query = f'DELETE from collection_point WHERE id={} AND vo_id={} and participant_id={};'
        # for cell in row:
        #     print(cell.value)
        row_data = [cell.value for cell in row]
        all_data.append(row_data)
    break

print(all_data)

for data in all_data:
    # Delete from enterprise where id = '0d7f709d-571b-4243-8b17-bfbe4c0f8b92';
    query_list.append(f"Delete from enterprise where id='{data[0]}';")

for query in query_list:
    print(query)

print(len(query_list))