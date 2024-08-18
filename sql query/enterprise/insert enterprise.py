"""SELECT * from enterprise WHERE cohort_id=parent_cohort_id ALLOW FILTERING; """
"""copy & paste on xl sheet & run the script"""

import openpyxl
import uuid
import csv
import requests
import json
import sys

print('Welcome to UPG programme configuration DB!')
print('***** All Environments *****:')
environment = {
    '1': ['Stage', 'https://upgapstg.brac.net'],
    '2': ['Training', 'https://trainingupg.brac.net'],
    '3': ['Production', 'https://upgbd.brac.net']
}

for env in environment:
    print(env, environment[env][0])

# print('$$$ Select Environment:')
selected_environment = str(input('$$$ Select Environment: '))
base_url = ''
credential = {"email": "admin@brac.net", "password": "123456"}
for env in environment:
    if environment[selected_environment][0] == environment[env][0]:
        base_url = environment[env][1]
        if environment[env][0].upper() == 'PRODUCTION':
            credential = {"email": "admin@brac.net", "password": "12345@#"}
        break

# getting access token by login
login_json = requests.post(f'{base_url}/upg-auth/api/v1/account/login',
                          data = credential)

if login_json.status_code == 200:
    print('Success! Logged in successfully!!')
    # deserialize a JSON formatted string into a Python object
    login_data = json.loads(login_json.content)

    access_token = login_data['result']['access_token']

    # Fetch program and cohort data in a single API call
    program_data = requests.get(f'{base_url}/upg-participant-selection/api/v1/program',
                                headers={'Authorization': f"Bearer {access_token}"})
    program_info = json.loads(program_data.content)
    all_program = program_info['resultset']

    # Initialize program dictionary
    program_dictionary = []

    # Iterate over programs
    for index, program in enumerate(all_program):
        if program['is_active']:
            # Initialize cohorts list for each program
            cohorts_data = []
            # Fetch cohorts data for the current program
            cohorts_of_program_json = requests.get(
                f'{base_url}/upg-participant-selection/api/v1/cohort/{program["id"]}',
                headers={'Authorization': f"Bearer {access_token}"})
            cohorts_of_program_info = json.loads(cohorts_of_program_json.content)
            cohorts_values = cohorts_of_program_info['resultset']
            # Iterate over cohorts of the program
            cohort_serial = 0
            for single_cohort in cohorts_values:
                if single_cohort['is_active']:
                    cohort_serial += 1
                    cohorts_data.append({
                        'cohort_serial': cohort_serial,
                        'cohort_name': single_cohort['cohort'],
                        'cohort_id': single_cohort['id']
                    })

            # Add program info with cohorts to program_dictionary
            program_dictionary.append({
                'Serial': index + 1,
                'Program_name': program['program_name'],
                'Program_id': program['id'],
                'Cohorts': cohorts_data
            })

    # Print all programs
    print('***** All Programs *****:')
    for program in program_dictionary:
        print(program['Serial'], program['Program_name'])

    # Select parent program and cohort
    selected_parent_program = int(input('$$$$$ Select Parent Program: '))
    selected_parent_program_info = program_dictionary[selected_parent_program - 1]  # Adjusting index
    print(f"Select Cohort of {selected_parent_program_info['Program_name']}: ")

    for cohort in selected_parent_program_info['Cohorts']:
        print(cohort['cohort_serial'], cohort['cohort_name'])
    selected_parent_cohort = int(input('*** Enter Parent cohort: '))

    # Get selected parent cohort info
    selected_parent_cohort_info = selected_parent_program_info['Cohorts'][
        selected_parent_cohort - 1]  # Adjusting index
    parent = selected_parent_program_info['Program_name'] + ' ' + selected_parent_cohort_info['cohort_name']

    ##### Select child program and cohort
    print("*****Now, select Child program and cohort*****")
    print('All Programs:')
    for program in program_dictionary:
        print(program['Serial'], program['Program_name'])

    selected_child_program = int(input('$$$$$ Select Child Program: '))
    selected_child_program_info = program_dictionary[selected_child_program - 1]  # Adjusting index
    print(f"Select Cohort of {selected_child_program_info['Program_name']}: ")
    for cohort in selected_child_program_info['Cohorts']:
        print(cohort['cohort_serial'], cohort['cohort_name'])

    selected_child_cohort = int(input('Enter child cohort: '))
    # Get selected child cohort info
    selected_child_cohort_info = selected_child_program_info['Cohorts'][selected_child_cohort - 1]  # Adjusting index
    child = selected_child_program_info['Program_name'] + " " + selected_child_cohort_info['cohort_name']

    if parent == child:
        print("Program & cohort of parent and child is similar, can't advance further.")
        sys.exit(5)

    child_program_name = selected_child_program_info['Program_name']
    child_program_id = selected_child_program_info['Program_id']
    child_cohort_name = selected_child_cohort_info['cohort_name']
    child_cohort_id = selected_child_cohort_info['cohort_id']

    parent_program_name = selected_parent_program_info['Program_name']
    parent_program_id = selected_parent_program_info['Program_id']
    parent_cohort_name = selected_parent_cohort_info['cohort_name']
    parent_cohort_id = selected_parent_cohort_info['cohort_id']

    """get enterprise category for both parent & child program by cohort_id"""
    all_enterprise_category_json = requests.get(
        f"{base_url}/upg-participant-selection/api/v1/enterprise-category/all",
        headers={'Authorization': f"Bearer {access_token}"})

    all_enterprise_category_data = json.loads(all_enterprise_category_json.content)
    all_enterprise_category_info = all_enterprise_category_data['resultset']

    parent_category_map = {}
    child_category_map = {}
    for enterprise_category in all_enterprise_category_info:
        if enterprise_category['cohort_id'] == child_cohort_id:
            child_category_map[enterprise_category['name']] = enterprise_category['id']
        if enterprise_category['cohort_id'] == parent_cohort_id:
            parent_category_map[enterprise_category['name']] = enterprise_category['id']

    final_category_map = {value: child_category_map[key.rstrip('.')] for key, value in parent_category_map.items()}

    # Load the workbook
    workbook = openpyxl.load_workbook('../Book1.xlsx')

    # Get all sheet names
    sheet_names = workbook.sheetnames

    all_data = []
    query_list = []

    for sheet_name in sheet_names:
        # Select the sheet by name
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column):
            row_data = [cell.value for cell in row]
            all_data.append(row_data)
        break

    # print(all_data)
    uuid_values = []

    for data in all_data:
        uuid_value = uuid.uuid1()
        while uuid_value in uuid_values:
            uuid_value = uuid.uuid1()
        uuid_values.append(uuid_value)

        if data[1] in final_category_map:
            category_id = final_category_map[data[1]]

            query_list.append(f'Insert into enterprise ("id","category_id","client_created_at","client_updated_at","cohort_id","created_at","created_by","description","enterprise_code","is_active","main_asset_id","main_asset_quantity","name","supporting_asset_id","supporting_asset_quantity","updated_at","updated_by") VALUES '
            f"('{uuid_value}', {category_id}, '{data[2]}', '{data[3]}', {child_cohort_id}, toTimestamp(now()), {data[6]}, {data[7]}, '{data[8]}', {str(data[9]).lower()}, {data[10]}, {data[11]}, '{data[12]}', {data[13]}, {data[14]}, toTimestamp(now()), {data[16]});")

    updated_queries = [query.replace("None", '0') for query in query_list]

    for query in updated_queries:
        print(query)

    print(len(query_list))