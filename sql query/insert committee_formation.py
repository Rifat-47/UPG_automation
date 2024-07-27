"""INSERT INTO "upg"."committee_formation" ("cohort_id","id","created_at","created_by",
"designation_id","designation_name","is_active","max_committee_member","min_committee_member",
"updated_at","updated_by") VALUES (df90c5c9-7fb3-4633-99db-259b2ce82990,1974c166-1cf9-4457-967a-60a92cad9d6c,
'2023-11-05 12:29:34.301+0000','cac6aebd-d346-4112-ab7c-b2701f66d90e','43793e5f-7ae9-4403-a6b2-f9652ebf533a',
'সাধারণ সদস্য','true',100,100,'2023-11-05 12:29:34.301+0000',NULL);"""

"""DELETE from committee_formation WHERE cohort_id=6a1325d4-07cc-4985-b7cc-c945c4f536fe;"""

import openpyxl

# Load the workbook
workbook = openpyxl.load_workbook('Book1.xlsx')

# Get all sheet names
sheet_names = workbook.sheetnames

all_data = []
query_list = []

for sheet_name in sheet_names:
    # Select the sheet by name
    sheet = workbook[sheet_name]
    # print(sheet)
    # print(sheet.min_row, sheet.max_row)
    # print(sheet.min_column, sheet.max_column)

    for row in sheet.iter_rows(min_row=sheet.min_row, max_row=sheet.max_row, min_col=sheet.min_column, max_col=sheet.max_column):
        # query = f'DELETE from collection_point WHERE id={} AND vo_id={} and participant_id={};'
        # for cell in row:
        #     print(cell.value)
        row_data = [cell.value for cell in row]
        all_data.append(row_data)
    break

print(all_data)

for data in all_data:
    query_list.append(f'INSERT INTO committee_formation ("cohort_id","id","created_at","created_by","designation_id","designation_name","is_active","max_committee_member","min_committee_member","updated_at","updated_by") VALUES '
                      f"(6a1325d4-07cc-4985-b7cc-c945c4f536fe, uuid(), toTimestamp(now()), '{data[3]}', '{data[4]}', '{data[5]}', {str(data[6]).lower()}, {data[7]}, {data[8]}, "
                      f"toTimestamp(now()), {data[10]});")

for query in query_list:
    print(query)

print(len(query_list))

